"""
rm_efficiency_report.py

Builds a per-RM (relationship manager) performance report from raw
transaction-level Excel data, consolidating duplicate rows per RM/month and
computing an Efficiency metric (Retention / Overall_Salary) to flag which
RMs are generating retention value relative to their cost.

The output Excel report is color-coded:
    - Red   : Efficiency <= 0.00 (poor)
    - Yellow: 0.00 < Efficiency <= 1.99 (average)
    - Green : Efficiency > 1.99 (good)
  and any metric that declined month-over-month is highlighted in red.

Usage:
    python rm_efficiency_report.py --input path/to/input.xlsx --output path/to/report.xlsx
"""

import argparse
import os
import re
import shutil
import sys
import tempfile
import time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REQUIRED_COLUMNS = [
    "emp_id", "rm_code", "rm_name", "d_branch", "d_region", "month_period",
    "Active POS", "NOP", "Gross Premium", "Payin", "Payout", "Retention", "Overall_Salary"
]

METRICS = ["Active POS", "NOP", "Gross Premium", "Payin", "Payout", "Retention", "Overall_Salary"]


def generate_rm_efficiency_report(input_file: str, output_file: str) -> bool:
    """
    Generate a performance report for relationship managers by transforming
    raw Excel data. Groups data by rm_code to provide consolidated metrics
    per RM, eliminating duplicates in the final report.
    """
    try:
        df = pd.read_excel(input_file)
        print(f"Loaded data from {input_file}")
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        return False

    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        print(f"Missing required columns: {', '.join(missing_columns)}")
        return False

    # For each unique rm_code, pick the most common name/region/branch/emp_id
    rm_attributes = {}
    for attribute in ['emp_id', 'rm_name', 'd_region', 'd_branch']:
        rm_attributes[attribute] = df.groupby('rm_code')[attribute].agg(
            lambda x: x.value_counts().index[0] if len(x.value_counts()) > 0 else "Unknown"
        ).to_dict()

    metrics_summed = df.groupby(['rm_code', 'month_period']).agg({
        "Active POS": "sum",
        "NOP": "sum",
        "Gross Premium": "sum",
        "Payin": "sum",
        "Payout": "sum",
        "Retention": "sum",
        "Overall_Salary": "sum"
    }).reset_index()

    for attribute, mapping in rm_attributes.items():
        metrics_summed[attribute] = metrics_summed['rm_code'].map(mapping)

    duplicate_check = metrics_summed.groupby(['rm_code', 'month_period']).size()
    if (duplicate_check > 1).any():
        print("Warning: duplicates found after grouping, deduplicating.")
        metrics_summed = metrics_summed.drop_duplicates(['rm_code', 'month_period'])

    metrics_summed = metrics_summed.sort_values(['d_region', 'd_branch', 'rm_code', 'month_period'])

    df_melted = metrics_summed.melt(
        id_vars=["d_region", "d_branch", "rm_code", "rm_name", "emp_id", "month_period"],
        value_vars=METRICS,
        var_name="Metric",
        value_name="Value"
    )
    df_melted["Metric_Month"] = df_melted["Metric"] + " - " + df_melted["month_period"].astype(str)

    pivot_df = df_melted.pivot_table(
        index=["d_region", "d_branch", "rm_code", "rm_name", "emp_id"],
        columns="Metric_Month",
        values="Value",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    duplicate_count = pivot_df.duplicated(['rm_code']).sum()
    if duplicate_count > 0:
        print(f"Warning: {duplicate_count} duplicate RM codes found, removing.")
        pivot_df = pivot_df.drop_duplicates(['rm_code'])

    pivot_df.columns.name = None
    pivot_df.columns = [str(col) for col in pivot_df.columns]

    desired_metric_order = METRICS + ["Efficiency"]
    id_cols = ["d_region", "d_branch", "rm_code", "rm_name", "emp_id"]
    all_months = sorted(set(col.split(' - ')[1] for col in pivot_df.columns if ' - ' in col))

    # Efficiency = Retention / Overall_Salary, per month
    for month in all_months:
        retention_col = f"Retention - {month}"
        salary_col = f"Overall_Salary - {month}"
        efficiency_col = f"Efficiency - {month}"
        if retention_col in pivot_df.columns and salary_col in pivot_df.columns:
            pivot_df[efficiency_col] = pivot_df.apply(
                lambda row: row[retention_col] / row[salary_col] if row[salary_col] != 0 else 0,
                axis=1
            ).round(2)

    ordered_columns = id_cols.copy()
    for metric in desired_metric_order:
        for month in all_months:
            col_name = f"{metric} - {month}"
            if col_name in pivot_df.columns:
                ordered_columns.append(col_name)
    pivot_df = pivot_df[ordered_columns]

    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    try:
        temp_dir = tempfile.mkdtemp()
        temp_file = os.path.join(temp_dir, f"temp_report_{int(time.time())}.xlsx")
        pivot_df.to_excel(temp_file, index=False)
        apply_excel_formatting(temp_file, pivot_df)
        try:
            shutil.copy2(temp_file, output_file)
            print(f"Report generated: {output_file}")
            return True
        except Exception as e:
            print(f"Error copying to final destination: {e}")
            print(f"Report is available at the temporary location: {temp_file}")
            return False
        finally:
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass
    except Exception as e:
        print(f"Error writing the Excel report: {e}")
        return False


def apply_excel_formatting(excel_file, df):
    """Merge metric header cells, highlight decline/efficiency, and add borders."""
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    header_font = Font(name='Arial', size=11, bold=True, color='000000')
    header_fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    decline_fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')

    efficiency_poor_fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
    efficiency_good_fill = PatternFill(start_color='B5E6A2', end_color='B5E6A2', fill_type='solid')
    efficiency_average_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    worksheet.insert_rows(1)

    metrics = METRICS + ["Efficiency"]
    metric_columns = {metric: [] for metric in metrics}
    id_columns = 5
    metric_month_columns = {}
    pattern = re.compile(r"(.*) - (.*)")

    for col_idx, col_name in enumerate(df.columns):
        col_letter = get_column_letter(col_idx + 1)
        if col_idx < id_columns:
            continue
        match = pattern.match(col_name)
        if match:
            metric, month = match.groups()
            if metric in metrics:
                metric_columns[metric].append(col_letter)
            metric_month_columns.setdefault(metric, {})[month] = col_letter

    for metric, columns in metric_columns.items():
        if not columns:
            continue
        start_col, end_col = columns[0], columns[-1]
        worksheet.merge_cells(f"{start_col}1:{end_col}1")
        cell = worksheet[f"{start_col}1"]
        cell.value = metric
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

        last_col_letter = columns[-1]
        for row in range(1, worksheet.max_row + 1):
            cell = worksheet[f"{last_col_letter}{row}"]
            current_border = cell.border if cell.border else Border()
            cell.border = Border(
                left=current_border.left or Side(style='thin', color='000000'),
                right=Side(style='thick', color='000000'),
                top=current_border.top or Side(style='thin', color='000000'),
                bottom=current_border.bottom or Side(style='thin', color='000000')
            )

    for col in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col)
        cell = worksheet[f"{col_letter}2"]
        if col > id_columns:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    for row in range(3, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            cell = worksheet[f"{col_letter}{row}"]
            if not (cell.border and cell.border.right and cell.border.right.style == 'thick'):
                cell.border = thin_border

    for col in range(1, id_columns + 1):
        col_letter = get_column_letter(col)
        worksheet[f"{col_letter}2"].border = thin_border

    months = sorted(set(m for mm in metric_month_columns.values() for m in mm.keys()))
    try:
        month_order = {}
        for i, month in enumerate(months):
            if month.lower() == 'april':
                month_order[month] = 4
            elif month.isdigit():
                month_order[month] = int(month)
            else:
                try:
                    month_order[month] = pd.to_datetime(month, format='%B').month
                except Exception:
                    month_order[month] = i
        months = sorted(months, key=lambda m: month_order[m])
    except Exception as e:
        print(f"Warning: error sorting months: {e}. Using original order.")

    for row in range(3, worksheet.max_row + 1):
        for metric in metrics:
            if metric not in metric_month_columns:
                continue
            metric_months = metric_month_columns[metric]

            if metric == "Efficiency":
                for month in months:
                    if month not in metric_months:
                        continue
                    cell_ref = f"{metric_months[month]}{row}"
                    try:
                        efficiency_value = worksheet[cell_ref].value or 0
                        if efficiency_value <= 0.00:
                            worksheet[cell_ref].fill = efficiency_poor_fill
                        elif efficiency_value > 1.99:
                            worksheet[cell_ref].fill = efficiency_good_fill
                        else:
                            worksheet[cell_ref].fill = efficiency_average_fill
                    except Exception as e:
                        print(f"Warning: efficiency highlight error row {row}, month {month}: {e}")
                continue

            if len(metric_months) <= 1:
                continue

            for i in range(1, len(months)):
                current_month, prev_month = months[i], months[i - 1]
                if current_month not in metric_months or prev_month not in metric_months:
                    continue
                current_cell_ref = f"{metric_months[current_month]}{row}"
                prev_cell_ref = f"{metric_months[prev_month]}{row}"
                try:
                    current_value = worksheet[current_cell_ref].value or 0
                    prev_value = worksheet[prev_cell_ref].value or 0
                    if current_value < prev_value:
                        worksheet[current_cell_ref].fill = decline_fill
                except Exception as e:
                    print(f"Warning: decline compare error row {row}, metric {metric}: {e}")

    try:
        workbook.save(excel_file)
    finally:
        workbook.close()


def main():
    parser = argparse.ArgumentParser(
        description="Generate a per-RM efficiency report from raw transaction data."
    )
    parser.add_argument("--input", required=True, help="Path to input Excel file")
    parser.add_argument("--output", required=True, help="Path for output Excel report")
    args = parser.parse_args()

    success = generate_rm_efficiency_report(args.input, args.output)
    if success:
        print("\nReport generation complete.")
        print("Red   cells: declining metric vs previous month")
        print("Efficiency color coding -> Red: <= 0.00 | Yellow: 0.00-1.99 | Green: > 1.99")
    else:
        print("\nReport generation failed.")
        sys.exit(1)


if __name__ == "__main__":
    main()
