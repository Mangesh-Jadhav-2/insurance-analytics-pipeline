"""
weighted_average_calc.py

Calculates a segment-weighted premium average per branch and ranks branches
by their weighted performance. Segments are weighted differently (e.g. Health
counts more heavily than Motor) to reflect their relative contribution to
profitability, not just raw premium volume.

Outputs a formatted Excel workbook with one sheet per parameter (Premium,
Payin, Payout, Retention), each showing actual values alongside the
weighted/ranked view.

Usage:
    python weighted_average_calc.py --input path/to/input.xlsx --output path/to/output.xlsx
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------- CONFIG ---------------- #
PARAMS = ['Prem', 'PII', 'PO', 'Ret']
REQUIRED_COLUMNS = ['d_branch', 'segment'] + PARAMS

SEGMENT_MULTIPLIER = {
    'HEALTH': 4,
    'LIFE': 2,
    'MOTOR': 1,
    'SME': 2
}

SEGMENT_ORDER = ["HEALTH", "LIFE", "MOTOR", "SME"]
# ---------------------------------------- #


def validate_input_file(file_path):
    """Validate that input file exists and is readable."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")
    if not path.suffix.lower() in ['.xlsx', '.xls']:
        raise ValueError("Input file must be Excel format (.xlsx or .xls)")
    return path


def validate_dataframe(df):
    """Validate that dataframe has required columns and data."""
    missing_cols = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns: {missing_cols}")

    if df.empty:
        raise ValueError("Input dataframe is empty")

    if df['d_branch'].isna().all():
        raise ValueError("No branch data found")
    if df['segment'].isna().all():
        raise ValueError("No segment data found")

    return True


def create_output_directory(file_path):
    """Ensure output directory exists."""
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def process_weighted_average(df, param):
    """Calculate weighted average for a given parameter."""
    actual = pd.pivot_table(
        df,
        index='d_branch',
        columns='segment',
        values=param,
        aggfunc='sum',
        fill_value=0
    )

    for seg in SEGMENT_ORDER:
        if seg not in actual.columns:
            actual[seg] = 0

    actual = actual[SEGMENT_ORDER]

    weighted = actual.copy()
    for seg, mult in SEGMENT_MULTIPLIER.items():
        if seg in weighted.columns:
            weighted[seg] *= mult

    weighted['Total'] = weighted.sum(axis=1)
    weighted = weighted.sort_values('Total', ascending=False)
    weighted['Rank'] = range(1, len(weighted) + 1)

    actual = actual.loc[weighted.index]

    actual.columns = pd.MultiIndex.from_product([["Gross_Premium"], actual.columns])
    weighted.columns = pd.MultiIndex.from_product(
        [["Premium Weighted Average"], weighted.columns]
    )

    final_df = pd.concat([actual, weighted], axis=1)
    return final_df


def apply_excel_formatting(workbook):
    """Apply formatting to all worksheets."""
    blue = PatternFill("solid", fgColor="CFE8F3")
    yellow = PatternFill("solid", fgColor="FFD966")
    green = PatternFill("solid", fgColor="92D050")

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for ws in workbook.worksheets:
        ws.freeze_panes = "B3"

        ws.cell(row=1, column=2).value = "Gross_Premium"
        ws.cell(row=1, column=6).value = "Premium Weighted Average"

        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
        ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=11)

        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin

        for col in range(2, 6):  # Gross Premium columns
            ws.cell(row=2, column=col).fill = blue

        for col in range(6, 11):  # Weighted columns
            ws.cell(row=2, column=col).fill = yellow

        ws.cell(row=2, column=11).fill = green  # Rank column header only

        for col_idx in range(1, 13):  # Columns A through L
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15


def main():
    parser = argparse.ArgumentParser(
        description="Calculate segment-weighted branch premium averages."
    )
    parser.add_argument("--input", required=True, help="Path to input Excel file")
    parser.add_argument("--output", required=True, help="Path for output Excel report")
    args = parser.parse_args()

    try:
        print("Starting weighted average calculation...")

        input_path = validate_input_file(args.input)
        print(f"[OK] Input file validated: {input_path}")

        df = pd.read_excel(input_path)
        df['segment'] = df['segment'].str.upper()

        validate_dataframe(df)
        print(f"[OK] Data validated: {len(df)} rows, {len(df['d_branch'].unique())} branches")

        output_path = create_output_directory(args.output)
        print("[OK] Output directory ready")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for param in PARAMS:
                print(f"  Processing parameter: {param}")
                final_df = process_weighted_average(df, param)
                final_df.to_excel(writer, sheet_name=param, index=True)

        print("[OK] Data written to Excel")

        wb = load_workbook(output_path)
        apply_excel_formatting(wb)
        wb.save(output_path)

        print(f"\nDone. Output saved to:\n  {output_path}")

    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"Data validation error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {type(e).__name__}: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
