"""
branch_excel_splitter.py

Splits a single Excel workbook into one file per branch (based on the
`d_branch` column), preserving the original header styling, column widths,
and applying consistent color-coded header formatting and borders to each
output file. Useful for distributing branch-specific reports without
manually copy-pasting rows.

Usage:
    python branch_excel_splitter.py
    (then follow the prompts for input file and output folder)
"""

import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Header text -> (fill color, font color). Extend this to match your own
# report's column headers.
HEADER_COLORS = {
    'LOB': {'fill': 'FF156082', 'font': 'FFFFFFFF'},
    'd_branch': {'fill': 'FF156082', 'font': 'FFFFFFFF'},
    'PY Achieved': {'fill': 'FFF7C7AC', 'font': 'FF000000'},
    'YTD Achieved': {'fill': 'FFF7C7AC', 'font': 'FF000000'},
    'Target': {'fill': 'FFA6C9EC', 'font': 'FF000000'},
    'YTD - Fresh Achieved': {'fill': 'FFDAF2D0', 'font': 'FF000000'},
    'YTD - Renewals Achieved': {'fill': 'FFDAF2D0', 'font': 'FF000000'},
    'Total YTD Achieved': {'fill': 'FF83E28E', 'font': 'FF000000'},
    'YTD Variance': {'fill': 'FFFFFF99', 'font': 'FF000000'},
    '%': {'fill': 'FFA6C9EC', 'font': 'FF000000'},
}


def apply_header_formatting(ws):
    """Color-code header row 1 based on (fuzzy) matches against HEADER_COLORS."""
    for col_idx, cell in enumerate(ws[1], 1):
        header_text = cell.value
        stripped_header = str(header_text).strip() if header_text else ""

        matched = None
        if stripped_header in HEADER_COLORS:
            matched = HEADER_COLORS[stripped_header]
        else:
            for key, colors in HEADER_COLORS.items():
                if key.split()[0] in stripped_header:
                    matched = colors
                    break

        if matched:
            cell.fill = PatternFill(start_color=matched['fill'], end_color=matched['fill'], fill_type='solid')
            cell.font = Font(color=matched['font'], bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def apply_outer_border(ws, max_data_row):
    """Apply a thick outer border around the data range of the sheet."""
    max_col = ws.max_column
    for row in range(1, max_data_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            left = Side(style='thick', color='FF000000') if col == 1 else cell.border.left
            right = Side(style='thick', color='FF000000') if col == max_col else cell.border.right
            top = Side(style='thick', color='FF000000') if row == 1 else cell.border.top
            bottom = Side(style='thick', color='FF000000') if row == max_data_row else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def split_excel_by_branch(input_file, output_folder='output_files'):
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found!")
        return

    Path(output_folder).mkdir(parents=True, exist_ok=True)

    try:
        original_wb = load_workbook(input_file)
        original_ws = original_wb.active

        df = pd.read_excel(input_file)
        if 'd_branch' not in df.columns:
            print("Error: 'd_branch' column not found!")
            return

        unique_branches = df['d_branch'].unique()
        print(f"Found {len(unique_branches)} unique branches")

        for branch in unique_branches:
            branch_indices = df[df['d_branch'] == branch].index.tolist()
            safe_branch_name = str(branch).replace('/', '_').replace('\\', '_').replace(' ', '_')
            output_file = f"{output_folder}/{safe_branch_name}.xlsx"

            new_wb = load_workbook(input_file)
            new_ws = new_wb.active

            rows_to_delete = [
                row_idx for row_idx in range(2, original_ws.max_row + 1)
                if row_idx - 2 not in branch_indices
            ]
            for row_idx in reversed(rows_to_delete):
                new_ws.delete_rows(row_idx)

            max_data_row = len(branch_indices) + 1
            apply_header_formatting(new_ws)
            apply_outer_border(new_ws, max_data_row)

            new_wb.save(output_file)
            print(f"Created: {output_file} ({len(branch_indices)} rows)")

        print(f"Successfully created {len(unique_branches)} files")

    except Exception as e:
        print(f"Error: {str(e)}")


if __name__ == "__main__":
    input_file = input("Enter Excel file path: ").strip().strip('"').strip("'")
    output_folder = input("Output folder (press Enter for 'output_files'): ").strip() or 'output_files'
    split_excel_by_branch(input_file, output_folder)
